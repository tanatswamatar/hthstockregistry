from django.shortcuts import render, redirect 
from .forms import AllocationForm 
from django.db.models import Sum, Count, F, Q
from decimal import Decimal
from .models import Grower, FieldOfficer, InventoryItem, Allocation
from django.contrib import messages

#-----------------------------------------------------------------------------------------------------------------------
import os
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from .models import Grower, Allocation
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Grower, InventoryItem, Allocation
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Sum, F
from decimal import Decimal
from django.contrib.auth.decorators import login_required

@login_required
def export_growers_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grower Inputs Final"

    # Define Colors based on your image
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    # 1. SETUP HEADERS (Rows 1 and 2)
    items = list(InventoryItem.objects.all().order_by('id'))
    
    # Static First Columns
    ws.merge_cells('A1:A2')
    ws['A1'] = "GROWER NO."
    ws.merge_cells('B1:B2')
    ws['B1'] = "SURNAME"

    curr_col = 3
    # Dynamic Item Columns (Merged Header for Name, Split for Qty/Amount)
    for item in items:
        # Merge top cell for Item Name
        ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col+1)
        ws.cell(row=1, column=curr_col).value = item.name.upper()
        
        # Sub-headers
        ws.cell(row=2, column=curr_col).value = "QTY"
        ws.cell(row=2, column=curr_col+1).value = f"AMOUNT USD {item.unit_price}"
        curr_col += 2

    # Summary Columns Headers
    summary_headers = [
        ("SUB TOTAL", blue_fill),
        ("ADMIN 9%", orange_fill),
        ("TOTAL USD", blue_fill)
    ]

    for title, fill in summary_headers:
        ws.merge_cells(start_row=1, start_column=curr_col, end_row=2, end_column=curr_col)
        cell = ws.cell(row=1, column=curr_col)
        cell.value = title
        cell.fill = fill
        curr_col += 1

    # 2. FILL DATA
    growers = Grower.objects.all().prefetch_related('allocations', 'allocations__item')
    row_num = 3

    for grower in growers:
        ws.cell(row=row_num, column=1).value = grower.grower_no
        ws.cell(row=row_num, column=2).value = grower.surname
        
        col_idx = 3
        sub_total = Decimal('0.00')

        for item in items:
            alloc = grower.allocations.filter(item=item).aggregate(
                total_qty=Sum('quantity'),
                total_val=Sum(F('quantity') * F('item__unit_price'))
            )
            qty = alloc['total_qty'] or 0
            amt = alloc['total_val'] or Decimal('0.00')
            
            ws.cell(row=row_num, column=col_idx).value = qty
            ws.cell(row=row_num, column=col_idx+1).value = amt
            sub_total += amt
            col_idx += 2

        # Final Calculations
        admin_fee = sub_total * Decimal('0.09')
        grand_total = sub_total + admin_fee

        ws.cell(row=row_num, column=col_idx).value = sub_total
        ws.cell(row=row_num, column=col_idx+1).value = admin_fee
        ws.cell(row=row_num, column=col_idx+2).value = grand_total
        row_num += 1

    # 3. FINAL STYLING (Borders and Alignment)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=row_num-1, max_col=col_idx+2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row <= 2:
                cell.font = Font(bold=True)

    # 1. Increase the First 2 Columns (Grower No and Surname)
    ws.column_dimensions['A'].width = 20  # Grower No
    ws.column_dimensions['B'].width = 30  # Surname

    # 2. Increase the Last 3 Columns (Subtotal, Admin 9%, Total USD)
    # Since your columns are dynamic, we find the last column index
    last_col = ws.max_column

    # Helper to convert number index (like 15) to Excel letter (like 'O')
    from openpyxl.utils import get_column_letter

    for i in range(last_col - 2, last_col + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 25

    # Save and Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Heritage_Final_Sheet.xlsx"'
    wb.save(response)
    return response

@login_required
def generate_invoice_pdf(request, grower_id):
    # 1. Fetch Data from Database
    grower = get_object_or_404(Grower, id=grower_id)
    # Get today's allocations for this grower
    allocations = grower.allocations.filter(date_issued__date=datetime.date.today())

    # Prepare PDF Response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Heritage_Invoice_{grower.grower_no}.pdf"'

    # 2. Setup Canvas
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 40

    # ---------- LOGO (TOP RIGHT) ----------
    # Assumes logo is in your static folder: static/images/heritagelogo.png
    logo_path = os.path.join(settings.BASE_DIR,'growers','static', 'images', 'heritagelogo.png')
    if os.path.exists(logo_path):
        logo_width = 100
        logo_height = 100
        c.drawImage(logo_path, width - logo_width - 40, y - logo_height + 50,
                    width=logo_width, height=logo_height, preserveAspectRatio=True)

    # -------- HEADER (Your Exact Format) --------
    c.setFillColor(colors.green)
    c.setFont("Helvetica-Bold", 24)
    c.drawString(60, y, "HERITAGE")
    y -= 15
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(40, y, "Tobacco Handlers")
    y -= 25

    # -------- COMPANY INFO --------
    c.setFont("Helvetica", 9)
    c.drawString(40, y, "4350/1 Haydon Road")
    y -= 12
    c.drawString(40, y, "Mt Hampden, Harare")
    y -= 12
    c.drawString(40, y, "Phone: (263) 787719059 / 772807945")
    y -= 12
    c.drawString(40, y, "Email: rufaro@atlantistobacco.co.zw")
    y -= 25

    c.setFont("Helvetica", 10)
    c.drawString(40, y, f"Date: {datetime.datetime.now().strftime('%d/%m/%Y')}")
    y -= 12

    # Get DN from first allocation
    innum = 29910000
    invoice_no = allocations.first().delivery_note_no if allocations.exists() else str(innum) 
    
    c.drawString(40, y, f"Invoice Number: {invoice_no}")
    y -= 20

    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(width / 2, y, "Acknowledgement of receipt of Inputs for the season 2026 to 2027")
    y -= 80

    # -------- BILL TO (Mapping Database Fields) --------
    bill_y = height - 90
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(width - 40, bill_y, "Bill To:")
    c.setFont("Helvetica", 9)
    c.drawRightString(width - 40, bill_y - 15, f"{grower.grower_no}")
    c.drawRightString(width - 40, bill_y - 30, f"{grower.first_name} {grower.surname}")
    c.drawRightString(width - 40, bill_y - 45, f"{grower.farm}")
    c.drawRightString(width - 40, bill_y - 60, f"{grower.area}")

    # -------- TABLE DATA --------
    table_data = [["Quantity", "Description", "Unit Price ($)", "Net Price ($)"]]
    total_calc = 0

    for alloc in allocations:
        net_price = alloc.quantity * alloc.item.unit_price
        total_calc += net_price
        table_data.append([
            f"{alloc.quantity}",
            alloc.item.name,
            f"{alloc.item.unit_price:,.2f}",
            f"{net_price:,.2f}"
        ])

    table_data.append(["", "", "Subtotal", f"{total_calc:,.2f}"])
    table_data.append(["", "", "Additional Discount", "0%"])
    table_data.append(["", "", "Balance Due", f"{total_calc:,.2f}"])

    # -------- TABLE STYLING --------
    table = Table(table_data, colWidths=[60, 200, 120, 90])
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
        ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONT", (0,-3), (-1,-1), "Helvetica-Bold"),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))

    table.wrapOn(c, width, height)
    # Correctly calculate Y position for the table
    table_y = y - (len(table_data) * 18)
    table.drawOn(c, 40, table_y)

    # -------- FOOTER SIGNATURES --------
    y = table_y - 40
    c.setFont("Helvetica", 10)
    c.drawString(40, y, "Received in Good Order by:")
    y -= 20
    c.drawString(40, y, "Name:.........................................................................................")
    y -= 20
    c.drawString(40, y, "Signature: ....................................................................................")

    c.showPage()
    c.save()
    return response


@login_required
def dashboard(request):
    # 1. HANDLE SEARCH LOGIC
    query = request.GET.get('q')
    search_results = None
    if query:
        search_results = Grower.objects.filter(
            Q(grower_no__icontains=query) | Q(surname__icontains=query)
        )[:10]

    # 2. CALCULATE TOP SUMMARY STATS
    # Get all allocations to calculate total value
    all_allocations = Allocation.objects.all()
    total_value = all_allocations.aggregate(
        total=Sum(F('quantity') * F('item__unit_price'))
    )['total'] or 0
    
    admin_fee = float(total_value) * 0.09
    grand_total = float(total_value) + admin_fee

    # 3. FIELD OFFICER PERFORMANCE TABLE
    # This sums up the debt for every grower assigned to each officer
    officer_stats = FieldOfficer.objects.annotate(
        grower_count=Count('growers'),
        total_debt=Sum(F('growers__allocations__quantity') * F('growers__allocations__item__unit_price'))
    )

    # 4. WAREHOUSE STOCK WINDOW (The "Second Edit")
    # We use annotate to calculate 'total_issued' on the fly
    inventory = InventoryItem.objects.annotate(
        total_issued=Sum('allocation__quantity')
    ).order_by('name')

    # 5. ORGANIZE CONTEXT FOR THE TEMPLATE
    context = {
        'query': query,
        'search_results': search_results,
        'total_growers': Grower.objects.count(),
        'total_value': total_value,
        'admin_fee': admin_fee,
        'grand_total': grand_total,
        'officer_stats': officer_stats,
        'inventory': inventory,
    }
    
    return render(request, 'growers/dashboard.html', context)
@login_required
def officer_growers(request, officer_id):
    officer = FieldOfficer.objects.get(id=officer_id)
    growers = officer.growers.all().order_by('surname')
    return render(request, 'growers/officer_growers.html', {'officer': officer, 'growers': growers})



from django.db import transaction, models
from django.contrib import messages
from decimal import Decimal

@login_required
def allocate_stock(request):
    items = InventoryItem.objects.all()
    form = AllocationForm()

    if request.method == 'POST':
        form = AllocationForm(request.POST)
        if form.is_valid():
            g_no = form.cleaned_data['grower_no']
            dn_no = form.cleaned_data['delivery_note']

            try:
                grower = Grower.objects.get(grower_no=g_no)

                # --- STEP 1: VALIDATION PRE-CHECK ---
                # Check ALL items first to see if any would go negative
                for item in items:
                    qty_val = request.POST.get(f'qty_{item.id}')
                    if qty_val and float(qty_val) > 0:
                        qty = Decimal(qty_val)
                        if qty > item.current_stock:
                            messages.error(request, f"Insufficient stock for {item.name}! (Available: {item.current_stock}, Requested: {qty})")
                            return render(request, 'growers/allocate.html', {'form': form, 'items': items})

                # --- STEP 2: SAVE DATA (Atomic) ---
                has_entries = False
                with transaction.atomic():  # Ensures all or nothing is saved
                    for item in items:
                        qty_val = request.POST.get(f'qty_{item.id}')

                        if qty_val and float(qty_val) > 0:
                            qty = Decimal(qty_val)

                            # 1. Create Allocation
                            Allocation.objects.create(
                                grower=grower,
                                item=item,
                                quantity=qty,
                                delivery_note_no=dn_no
                            )

                            # 2. Update Stock Level
                            item.current_stock -= qty
                            item.save()
                            has_entries = True

                if has_entries:
                    messages.success(request, f"Successfully allocated stock to {grower.first_name} {grower.surname}")
                    return redirect('dashboard')
                else:
                    messages.warning(request, "No quantities were entered.")

            except Grower.DoesNotExist:
                messages.error(request, "Grower not found.")

    return render(request, 'growers/allocate.html', {'form': form, 'items': items})

@login_required
def grower_detail(request, grower_id):
    grower = Grower.objects.get(id=grower_id)
    # This grabs all allocations belonging to THIS specific grower
    personal_inventory = grower.allocations.all().order_by('-date_issued')

    # Calculate their specific total debt
    total_debt = sum(a.quantity * a.item.unit_price for a in personal_inventory)

    return render(request, 'growers/grower_detail.html', {
        'grower': grower,
        'allocations': personal_inventory,
        'total_debt': total_debt
    })


from django.shortcuts import render, redirect, get_object_or_404
from .forms import WageRequestForm
from .models import Grower
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

def wage_request(request):
    if request.method == 'POST':
        form = WageRequestForm(request.POST)
        if form.is_valid():
            # 1. Collect the data
            g_no = form.cleaned_data['grower_no']
            amount = form.cleaned_data['amount']
            desc = form.cleaned_data['description']

            # 2. Get the Grower object (we know it exists because of form validation)
            grower = Grower.objects.get(grower_no=g_no)

            # 3. Create a file buffer (like a temporary file in memory)
            buffer = io.BytesIO()

            # 4. Create the PDF object, using the buffer as its "file"
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            y = height - 20

            # --- DRAWING THE PDF CONTENT ---

            # ---------- LOGO (TOP RIGHT) ----------
            # Assumes logo is in your static folder: static/images/heritagelogo.png
            logo_path = os.path.join(settings.BASE_DIR,'growers','static', 'images', 'heritagelogo.png')
            if os.path.exists(logo_path):
                logo_width = 100
                logo_height = 100
                p.drawImage(logo_path, 40, y - logo_height + 50,width=logo_width, height=logo_height, preserveAspectRatio=True)
            # Header
            # -------- HEADER (Your Exact Format) --------
            p.setFillColor(colors.green)
            p.setFont("Helvetica-Bold", 24)
            p.drawString(60, y-40, "HERITAGE")
            y -= 15
            p.setFillColor(colors.black)
            p.setFont("Helvetica-Bold", 20)
            p.drawString(40, y-40, "Tobacco Handlers")

            # -------- COMPANY INFO --------
            p.setFont("Helvetica", 9)
            p.drawString(width-180, y, "4350/1 Haydon Road")
            y -= 12
            p.drawString(width - 180, y, "Mt Hampden, Harare")
            y -= 12
            p.drawString(width - 180, y, "Phone: (263) 787 719 059")
            y -=12
            p.drawString(width - 180, y, "Phone: (263) 772 892 527")
            y -=12
            p.drawString(width - 180, y, "Phone: (263) 772 897 945")
            y -=12
            p.drawString(width - 180, y, "Email: rufaro@atlantistobacco.co.zw")
            y -= 25

            p.setFont("Helvetica", 10)
            p.drawString(width - 180, y, f"Date: {datetime.datetime.now().strftime('%d/%m/%Y')}")
            y -= 12

            #Payment Header
            p.setFont("Helvetica-Bold", 15)
            p.drawString(2.6*inch,y,"PAYMENT REQUISTION FORM")
            y-=8

            # Grower Details Section
            p.setStrokeColor("black")
            p.rect(35, y-165, 500, height-642, stroke=1, fill=0)
            g=y-12
            p.setFont("Helvetica-Bold", 15)
            p.drawString(40, g, "Grower Details:")
            y -= 30

            p.setFont("Helvetica", 12)
            # y-coordinates decrease as you go down the page
            p.drawString(40, y, f"Name: {grower.first_name} {grower.surname}")
            y-= 12
            p.drawString(40, y, f"Farm: {grower.farm}")
            y-=12
            p.drawString(40, y, f"Area: {grower.area}")
            y-=12
            p.drawString(40, y, f"Phone: {grower.phone}")
            y-=18

            # Bank Details Section
            p.setFont("Helvetica", 12)
            p.drawString(40, y, "Bank Details:")
            y-=12

            p.setFont("Helvetica", 12)
            p.drawString(40, y, f"Bank Name: {grower.bank_name or 'N/A'}")
            y-=12
            p.drawString(40, y, f"Branch: {grower.branch_name or 'N/A'} ({grower.branch_code or '-'})")
            y-=12
            p.drawString(40, y, f"Account No: {grower.account_number or 'N/A'}")
            y-=12
            p.drawString(40, y, f"Account Holder: {grower.account_holder or 'N/A'}")
            y-=18
            p.drawString(40 , y , "Confirm Bank Details:___________________________________________________")
            y-=35

            # Transaction Request Details
            p.setFont("Helvetica-Bold", 15)
            p.drawString(40,y, "Request Details:")
            y-=18

            p.setFont("Helvetica", 12)
            p.setStrokeColor("black")
            p.rect(35, y-80, 500, height-700, stroke=1, fill=0)
            p.drawString(40,y, f"Requested Amount: ${amount}")
            y-=25
            p.setStrokeColor("black")
            p.rect(35, y-80, 500, height-700, stroke=1, fill=0)
            p.setFont("Helvetica-Bold", 12)
            p.drawString(40, y, f"Description: {desc}")
            y-=68

            # Farmer Signature line
            p.setFont("Helvetica-Bold", 12)
            p.drawString(40,y, "Farmer:")
            p.drawString(4.5*inch,y,"Date:")
            y-=10
            p.drawString(40,y,"Signature")
            y-=25

            #Administrator/Accounting recommandation
            p.setFont("Helvetica-Bold", 15)
            p.drawString(40,y, "Production Manager's Recommendation:")
            y-=8

            p.setFont("Helvetica-Bold", 12)
            p.setStrokeColor("black")
            p.rect(35, y-115, 500, height-675, stroke=1, fill=0)
            y-=100
            p.drawString(40, y, "Production Manager:")
            p.drawString(4.5*inch,y-4,"Date:")
            u= y+12
            p.line(35,u,535,u)
            y-=10
            p.drawString(40,y,"Signature")

            y-=25

            #Production Manager's
            p.setFont("Helvetica-Bold", 15)
            p.drawString(40,y, "Administrator/Accountant's Recommendation:")
            y-=8

            p.setFont("Helvetica-Bold", 12)
            p.setStrokeColor("black")
            p.rect(35, y-100, 500, height-690, stroke=1, fill=0)
            y-=80
            p.drawString(40, y, "Clerk:")
            p.drawString(4.5*inch,y-4,"Date:")
            u= y+12
            p.line(35,u,535,u)
            y-=10
            p.drawString(40,y,"Signature")

            #Contract Manager's
            p.setFont("Helvetica-Bold", 12)
            p.setStrokeColor("black")
            p.rect(35, y-45, 500, height-760, stroke=1, fill=0)
            y-=25
            p.drawString(40, y, "Contract Manager:")
            y-=10
            p.drawString(40,y,"Signature")


            # 5. Close the PDF object cleanly
            p.showPage()
            p.save()

            # 6. FileResponse sets the Content-Disposition header so the browser knows it's a PDF
            buffer.seek(0)
            return FileResponse(buffer, as_attachment=False, filename=f'request_{grower.grower_no}.pdf')

    else:
        form = WageRequestForm()

    return render(request, 'growers/wage_request.html', {'form': form})
